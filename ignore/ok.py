import openpyxl as op
import json 

wb = op.load_workbook('ignore/Schedule.xlsx')
ws = wb['Table 1']
schedule = {}
grouplist = ['N1', 'N2', 'N3', 'N4', 'N5', 'N6', 'N7', 'N8', 'N9']
for group in grouplist:
    routine = []
    for row in ws.iter_rows(min_row = 121, max_row = 180):
        groups = row[6].value.split('+')
        # print(type(groups))
        # print(groups)
        if group in groups or "N1-N9" in groups:
            # print("==========================")
            data = {}
            for cell in row:
                titlenot = f'{cell.column_letter}3'
                title = ws[titlenot].value 
                # print(title, ": ", cell.value)
                data[title] = cell.value
            routine.append(data)
        schedule[group] = routine
print(schedule)
with open("schedule2.json", "w") as f:
    json.dump(schedule, f)


# print(dir(cell))
# print(dir(ws))